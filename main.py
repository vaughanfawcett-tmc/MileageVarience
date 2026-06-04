"""CLI: enrich a Milcap (or Trip Purpose) Excel report with trip reason classifications.

Usage:
    python main.py <input.xlsx> [--output enriched.xlsx] [--rules rules.yaml]
                   [--reason-col "Trip Reason"] [--claimed-col "Claimed Miles"]
                   [--expected-col "Expected Miles"] [--sheet 0]

The script:
  - Reads the input workbook
  - Classifies each row's trip reason
  - Adds columns: Classification, Variance %, Match Term, Rationale
  - Highlights non-Acceptable rows for quick review
  - Writes an enriched workbook
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from classifier import (
    ACCEPTABLE,
    NOT_ACCEPTABLE,
    POTENTIALLY_ACCEPTABLE,
    TripReasonClassifier,
    _variance_pct,
)


FILL_ACCEPTABLE = PatternFill("solid", fgColor="C6EFCE")
FILL_POTENTIAL = PatternFill("solid", fgColor="FFEB9C")
FILL_NOT_ACCEPTABLE = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lower = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    return None


def enrich(
    input_path: Path,
    output_path: Path,
    rules_path: Path,
    sheet: str | int = 0,
    reason_col: str | None = None,
    claimed_col: str | None = None,
    expected_col: str | None = None,
) -> dict:
    df = pd.read_excel(input_path, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]

    reason_col = reason_col or find_column(
        df, ["Trip Reason", "Reason", "Comments", "Driver Reason", "Trip Purpose"]
    )
    claimed_col = claimed_col or find_column(
        df, ["Claimed Miles", "Claimed Mileage", "Actual Miles", "Reported Miles"]
    )
    expected_col = expected_col or find_column(
        df, ["Expected Miles", "Expected Mileage", "Google Miles", "Google Maps Miles"]
    )

    if reason_col is None:
        raise SystemExit(
            f"Could not find a trip reason column in {input_path}. "
            f"Pass --reason-col explicitly. Columns: {list(df.columns)}"
        )

    classifier = TripReasonClassifier(rules_path)

    categories, matches, rationales, variances = [], [], [], []
    for _, row in df.iterrows():
        reason = row.get(reason_col)
        claimed = row.get(claimed_col) if claimed_col else None
        expected = row.get(expected_col) if expected_col else None
        result = classifier.classify(reason, claimed, expected)
        categories.append(result.category)
        matches.append(result.matched_term or "")
        rationales.append(result.rationale)
        variances.append(_variance_pct(claimed, expected))

    df["Classification"] = categories
    df["Variance %"] = [round(v, 1) if v is not None else None for v in variances]
    df["Match Term"] = matches
    df["Rationale"] = rationales

    df.to_excel(output_path, index=False)
    _apply_formatting(output_path)

    summary = {
        "total": len(df),
        ACCEPTABLE: int((df["Classification"] == ACCEPTABLE).sum()),
        POTENTIALLY_ACCEPTABLE: int(
            (df["Classification"] == POTENTIALLY_ACCEPTABLE).sum()
        ),
        NOT_ACCEPTABLE: int((df["Classification"] == NOT_ACCEPTABLE).sum()),
        "reason_col": reason_col,
        "claimed_col": claimed_col,
        "expected_col": expected_col,
    }
    return summary


def _apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    headers = {cell.value: cell.column for cell in ws[1]}
    class_col = headers.get("Classification")
    if class_col is None:
        wb.save(path)
        return

    fills = {
        ACCEPTABLE: FILL_ACCEPTABLE,
        POTENTIALLY_ACCEPTABLE: FILL_POTENTIAL,
        NOT_ACCEPTABLE: FILL_NOT_ACCEPTABLE,
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        value = row[class_col - 1].value
        fill = fills.get(value)
        if fill:
            row[class_col - 1].fill = fill

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    p = argparse.ArgumentParser(description="Classify Milcap trip reasons.")
    p.add_argument("input", type=Path, help="Input .xlsx report")
    p.add_argument("--output", type=Path, default=None)
    p.add_argument("--rules", type=Path, default=Path(__file__).parent / "rules.yaml")
    p.add_argument("--sheet", default=0)
    p.add_argument("--reason-col", default=None)
    p.add_argument("--claimed-col", default=None)
    p.add_argument("--expected-col", default=None)
    args = p.parse_args()

    output = args.output or args.input.with_name(args.input.stem + "_classified.xlsx")

    summary = enrich(
        input_path=args.input,
        output_path=output,
        rules_path=args.rules,
        sheet=args.sheet,
        reason_col=args.reason_col,
        claimed_col=args.claimed_col,
        expected_col=args.expected_col,
    )

    print(f"\nClassified {summary['total']} trips -> {output}")
    print(f"  Reason column:   {summary['reason_col']}")
    print(f"  Claimed column:  {summary['claimed_col']}")
    print(f"  Expected column: {summary['expected_col']}")
    print(f"\n  {ACCEPTABLE:24} {summary[ACCEPTABLE]:>5}")
    print(f"  {POTENTIALLY_ACCEPTABLE:24} {summary[POTENTIALLY_ACCEPTABLE]:>5}")
    print(f"  {NOT_ACCEPTABLE:24} {summary[NOT_ACCEPTABLE]:>5}")


if __name__ == "__main__":
    main()
