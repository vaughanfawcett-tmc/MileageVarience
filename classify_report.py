"""Classify a real Mileage Variance / Milcap export with the LLM classifier.

Handles the actual TMC export schema:
  - Multiple data sheets ("UK Tax Year", "Calendar Tax Year").
  - Real columns: vcReason (reason), BusinessMileage (claimed),
    SystemCalculatedMileage (expected/system), % Difference, Purpose,
    Excluded Parent Company, Column1 (existing keep/remove decision).
  - Messy + multilingual free text.

It de-duplicates reason strings, classifies the distinct set via the LLM
(OpenRouter, packed + concurrent), maps results back to every row, writes an
enriched workbook, and — if a keep/remove column is present — prints how the
3-way classification compares to that existing decision.

Usage:
    export OPENROUTER_API_KEY=sk-or-v1-...
    # Validate on a sample first (cheap):
    python classify_report.py "Mileage Variance Last 30 Days v2.xlsx" --limit 200
    # Full run (Haiku default):
    python classify_report.py "Mileage Variance Last 30 Days v2.xlsx"
    # Higher-quality model once validated:
    python classify_report.py "...xlsx" --model anthropic/claude-sonnet-4.6

    # No key yet? Estimate scale/cost only:
    python classify_report.py "...xlsx" --estimate
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import llm_classifier as llm
from llm_classifier import ACCEPTABLE, NOT_ACCEPTABLE, POTENTIALLY_ACCEPTABLE

DATA_SHEETS = ["UK Tax Year", "Calendar Tax Year"]

REASON_CANDIDATES = ["vcReason", "Trip Reason", "Reason", "Comments", "Driver Reason"]
CLAIMED_CANDIDATES = ["BusinessMileage", "Claimed Miles", "Actual Miles"]
EXPECTED_CANDIDATES = ["SystemCalculatedMileage", "Expected Miles", "Google Miles"]
DECISION_CANDIDATES = ["Column1", "Decision", "Keep/Remove"]
EXCLUDE_CANDIDATES = ["Excluded Parent Company"]

FILL = {
    ACCEPTABLE: PatternFill("solid", fgColor="C6EFCE"),
    POTENTIALLY_ACCEPTABLE: PatternFill("solid", fgColor="FFEB9C"),
    NOT_ACCEPTABLE: PatternFill("solid", fgColor="FFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")

_XML_ARTIFACT = re.compile(r"_x000D_|\r")


def find_column(cols, candidates):
    lower = {str(c).lower().strip(): c for c in cols}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    return None


def normalise_reason(value) -> str:
    if value is None or (isinstance(value, float) and value != value):
        return ""
    return _XML_ARTIFACT.sub(" ", str(value)).strip()


def load_rows(path) -> pd.DataFrame:
    """Concatenate the data sheets that exist, tagging each row with its sheet.

    `path` may be a filesystem path or a file-like object (e.g. an upload).
    """
    xl = pd.ExcelFile(path)
    sheets = [s for s in DATA_SHEETS if s in xl.sheet_names]
    if not sheets:
        # Fall back to the first sheet that has a reason column.
        for s in xl.sheet_names:
            if find_column(xl.parse(s, nrows=0).columns, REASON_CANDIDATES):
                sheets = [s]
                break
    if not sheets:
        raise SystemExit(f"No sheet with a reason column found in {path}")

    frames = []
    for s in sheets:
        df = xl.parse(s)
        df.columns = [str(c).strip() for c in df.columns]
        df["_sheet"] = s
        frames.append(df)
    return pd.concat(frames, ignore_index=True)


def classify(df: pd.DataFrame, args) -> dict:
    reason_col = args.reason_col or find_column(df.columns, REASON_CANDIDATES)
    if reason_col is None:
        raise SystemExit(f"No reason column. Columns: {list(df.columns)}")
    exclude_col = find_column(df.columns, EXCLUDE_CANDIDATES)

    df["_reason"] = df[reason_col].map(normalise_reason)

    work = df
    excluded = 0
    if exclude_col is not None:
        mask = df[exclude_col].astype(str).str.strip().str.upper() == "Y"
        excluded = int(mask.sum())
        work = df[~mask]

    if args.limit:
        # Stratified-ish sample: spread across the frame so we see variety.
        work = work.iloc[:: max(1, len(work) // args.limit)].head(args.limit)

    distinct = sorted({r for r in work["_reason"]})
    print(f"Rows to classify: {len(work)} ({excluded} excluded companies skipped)")
    print(f"Distinct reasons: {len(distinct)}")

    client = llm.make_client()
    results = llm.classify(
        client,
        distinct,
        model=args.model,
        batch_size=args.batch_size,
        max_workers=1 if args.sync else args.workers,
    )

    by_reason = dict(zip(distinct, results))
    df["Classification"] = df["_reason"].map(
        lambda r: by_reason[r].category if r in by_reason else ""
    )
    df["Rationale"] = df["_reason"].map(
        lambda r: by_reason[r].rationale if r in by_reason else ""
    )
    return {"reason_col": reason_col, "classified": work, "df": df}


def validate_against_decision(df: pd.DataFrame):
    """If an existing keep/remove column exists, show how it relates to us."""
    decision_col = find_column(df.columns, DECISION_CANDIDATES)
    classified = df[df["Classification"] != ""]
    if decision_col is None or classified.empty:
        return
    dec = classified[decision_col].astype(str).str.strip().str.lower()
    if not dec.isin(["keep", "remove"]).any():
        return
    print("\nClassification vs existing keep/remove decision:")
    crosstab = pd.crosstab(classified["Classification"], dec)
    print(crosstab.to_string())
    print(
        "\n(keep/remove is the existing decision — meaning unconfirmed. "
        "Treat this as a comparison, not an accuracy score.)"
    )


def spot_check(df: pd.DataFrame, n: int):
    """Print N classified rows side by side for eyeballing the rubric.

    Spread evenly across the three categories so you see variety, and show the
    existing keep/remove decision (if present) for comparison.
    """
    classified = df[df["Classification"] != ""]
    if classified.empty:
        return
    decision_col = find_column(df.columns, DECISION_CANDIDATES)
    cats = [ACCEPTABLE, POTENTIALLY_ACCEPTABLE, NOT_ACCEPTABLE]
    per = max(1, n // len(cats))

    picks = []
    for cat in cats:
        sub = classified[classified["Classification"] == cat]
        if sub.empty:
            continue
        step = max(1, len(sub) // per)
        picks.append(sub.iloc[::step].head(per))
    sample = pd.concat(picks) if picks else classified.head(n)

    print(f"\n=== SPOT CHECK ({len(sample)} rows) ===")
    for _, row in sample.iterrows():
        reason = row["_reason"] or "(blank)"
        if len(reason) > 100:
            reason = reason[:97] + "..."
        tag = row["Classification"]
        if decision_col is not None:
            tag += f"   [existing: {str(row[decision_col]).strip()}]"
        print(f"\n[{tag}]")
        print(f"  reason: {reason}")
        print(f"  why:    {row['Rationale']}")


def write_workbook(df: pd.DataFrame, output: Path):
    drop = [c for c in ("_reason", "_sheet") if c in df.columns]
    df.drop(columns=drop).to_excel(output, index=False)

    wb = load_workbook(output)
    ws = wb.active
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    headers = {cell.value: cell.column for cell in ws[1]}
    cls_col = headers.get("Classification")
    if cls_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            fill = FILL.get(row[cls_col - 1].value)
            if fill:
                row[cls_col - 1].fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output)


def estimate(df: pd.DataFrame, args):
    """Rough scale/cost estimate without calling the API (no key needed)."""
    reason_col = args.reason_col or find_column(df.columns, REASON_CANDIDATES)
    df["_reason"] = df[reason_col].map(normalise_reason)
    distinct = {r for r in df["_reason"]}
    b = args.batch_size
    chunks = -(-len(distinct) // b)  # ceil
    # Rubric (~750 tok) sent once per chunk; ~30 tok per reason in, ~40 out.
    in_tok = chunks * 750 + len(distinct) * 30
    out_tok = len(distinct) * 40
    rates = {  # OpenRouter pass-through list price, $ per 1M tokens (in, out).
        "anthropic/claude-opus-4.8": (5.0, 25.0),
        "anthropic/claude-sonnet-4.6": (3.0, 15.0),
        "anthropic/claude-haiku-4.5": (1.0, 5.0),
    }
    print(f"Total rows:        {len(df)}")
    print(f"Distinct reasons:  {len(distinct)}")
    print(f"Packed @ {b}/call: {chunks} requests")
    print("\nRough OpenRouter cost (no batch discount; packing amortises the rubric):")
    for model, (ri, ro) in rates.items():
        cost = (in_tok * ri + out_tok * ro) / 1e6
        print(f"  {model:30} ~${cost:6.2f}")


def main():
    p = argparse.ArgumentParser(description="Classify a Milcap variance export with Claude.")
    p.add_argument("input", type=Path)
    p.add_argument("--output", type=Path, default=None)
    p.add_argument("--model", default=llm.DEFAULT_MODEL)
    p.add_argument("--reason-col", default=None)
    p.add_argument("--limit", type=int, default=None, help="Classify only N rows (sample).")
    p.add_argument("--batch-size", type=int, default=llm.DEFAULT_BATCH_SIZE,
                   help="Reasons packed per request.")
    p.add_argument("--workers", type=int, default=llm.DEFAULT_WORKERS,
                   help="Concurrent requests.")
    p.add_argument("--sync", action="store_true", help="Run sequentially (1 worker; for debugging).")
    p.add_argument("--spot-check", type=int, default=0, metavar="N",
                   help="After classifying, print N sample rows (reason + category + rationale).")
    p.add_argument("--estimate", action="store_true", help="Estimate scale/cost only; no API.")
    args = p.parse_args()

    df = load_rows(args.input)

    if args.estimate:
        estimate(df, args)
        return

    out = args.output or args.input.with_name(args.input.stem + "_classified.xlsx")
    result = classify(df, args)
    df = result["df"]

    classified = df[df["Classification"] != ""]
    print("\nClassification summary:")
    for cat in (ACCEPTABLE, POTENTIALLY_ACCEPTABLE, NOT_ACCEPTABLE):
        print(f"  {cat:24} {int((classified['Classification'] == cat).sum()):>6}")

    validate_against_decision(df)
    if args.spot_check:
        spot_check(df, args.spot_check)
    write_workbook(df, out)
    print(f"\nWrote {out}")


if __name__ == "__main__":
    main()
